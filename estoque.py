from ast import Break
from gettext import find
from optparse import OptionContainer
from numpy import empty, ones
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
import pandas as pd
import time, asyncio, os.path
from csv import DictReader
from os import listdir, rename
from os.path import isfile, join
from category import Category
from preco import Preco
from selenium.webdriver.common.action_chains import ActionChains
from xml.etree import ElementTree as ET
import xmltodict
import openpyxl
from senhas import Senhas
import json
from openpyxl import load_workbook

# options = Options()

# chromeOptions = webdriver.ChromeOptions()

# chromeOptions.add_argument("user-data-dir=/home/anderson/perfil/")

# driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chromeOptions)

# actions = ActionChains(driver)

# url = "https://seller.shopee.com.br/portal/product/category"


# driver.get(url) 




# with open("xmldrop.xml") as xml_file:
#     data_dict = xmltodict.parse(xml_file.read())
#     print(len(data_dict['dados']['produto']))


# with open("11.xlsx") as xlsx:
#     data_dict = xmltodict.parse(xlsx)
#     for data in data_dict:
#         print(data)

# with open('11.xlsx', "r+") as f:
#     f.readlines

#     open_file = f.readlines()
#     get_data = []
#     variacao = []
#     estoque = []
#     for data in DictReader(open_file):
#         print(data)





# ws = wb.get_sheet_by_name('Sheet1')
# # print(sheet.cell(row="", column=5).value)
# list(ws.iter_rows())


# wb = load_workbook('11.xlsx')
# ws = wb.get_sheet_by_name('Sheet1')

# wb = load_workbook('11.xlsx')
# sheet = wb['Sheet1']
# corpus = []
# cells = sheet.iter_rows()
# # cells = list(e)
# for i in cells:
#     print(i.value)
#     corpus.append(i[0].value)


wb = openpyxl.load_workbook("11.xlsx")

sheet = wb.active

# intera
# for data in sheet["5"]:
#     print(data.value)
# print(sheet['5'][1].value)

#   # !!! Adjust file name !!!
# for ws in wb.worksheets:
#     # Iterate over the columns and rows, search for the text and replace
#     for row in ws.iter_rows(max_col=8, min_col=0, max_row=5, min_row=5,):
#         print(row)
#         for cell in row:
#             print(cell.value)



def atualiza(sku, id):
    print(sku)

list_dict = {}
# list_sku = []
def search_sku():
    for ws in wb.worksheets:
        for row in ws.iter_rows(max_col=8, min_col=0, max_row=30, min_row=5,):
            if row[4].value != None:
                # list_sku.append([row[0].value, row[4].value])
                list_dict[row[0].value] = row[4].value


search_sku()

for i, a in enumerate(sheet, 5):
    if (sheet[i][0].value == None):
        break
    print(i)
    print(sheet[i][0].value)

print(list_dict['21929282857'])
